from __future__ import annotations

from datetime import datetime
from functools import wraps

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render

from .models import ZucchettiEmployee, NewRequest, PlantSetting


def zucchetti_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request: HttpRequest, *args, **kwargs):
        if not (request.user.groups.filter(name="zucchetti").exists() or request.user.is_superuser):
            return HttpResponse("Access denied.", status=403)
        return view_func(request, *args, **kwargs)
    return wrapper


# ── Import from Zucchetti Excel ───────────────────────────────────────────────

def _parse_zucchetti_excel(file_obj) -> list[ZucchettiEmployee]:
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        emp_code, subject_code, surname, name, card_code = (list(row) + [None] * 5)[:5]
        records.append(ZucchettiEmployee(
            emp_code=str(emp_code).strip(),
            subject_code=str(subject_code).strip() if subject_code else "",
            surname=str(surname).strip() if surname else "",
            name=str(name).strip() if name else "",
            card_code=str(card_code).strip() if card_code else "",
        ))
    return records


# ── New Request Excel parser ──────────────────────────────────────────────────

def _parse_request_excel(file_obj) -> list[NewRequest]:
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_code, name, plant, date_val, time_in, time_out = (list(row) + [None] * 6)[:6]
        if not emp_code:
            continue
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        import datetime as dt
        records.append(NewRequest(
            emp_code=str(emp_code).strip(),
            name=str(name).strip() if name else "",
            plant=str(plant).strip() if plant else "",
            date=date_val,
            time_in=time_in if isinstance(time_in, dt.time) else None,
            time_out=time_out if isinstance(time_out, dt.time) else None,
        ))
    return records


# ── DAT generation helpers ────────────────────────────────────────────────────

def _build_dat_preview() -> dict:
    plant_map = {ps.plant: ps.terminal for ps in PlantSetting.objects.all()}
    ok = skip_no_terminal = skip_no_time = skip_no_badge = 0
    missing_badges = []
    for row in NewRequest.objects.all():
        if not plant_map.get(row.plant):
            skip_no_terminal += 1
            continue
        if not row.time_in and not row.time_out:
            skip_no_time += 1
            continue
        emp_id = row.emp_code[1:].lstrip("0") or "0"
        emp_code_7 = emp_id.zfill(7)
        if not ZucchettiEmployee.objects.filter(emp_code=emp_code_7).exists():
            skip_no_badge += 1
            missing_badges.append(row.emp_code)
            continue
        ok += 1
    return {
        "ok": ok,
        "skip_no_terminal": skip_no_terminal,
        "skip_no_time": skip_no_time,
        "skip_no_badge": skip_no_badge,
        "missing_badges": missing_badges,
    }


# ── Views ─────────────────────────────────────────────────────────────────────

@zucchetti_required
def generate_dat_view(request: HttpRequest) -> HttpResponse:
    rows = NewRequest.objects.all().order_by("date", "time_in", "time_out")
    if not rows.exists():
        messages.error(request, "No request data — import Excel first.")
        return redirect("zucchetti_sync")

    plant_map = {ps.plant: ps.terminal for ps in PlantSetting.objects.all()}
    lines = []

    for row in rows:
        terminal = plant_map.get(row.plant)
        if not terminal:
            continue
        if row.time_in:
            direction, time_val = 0, row.time_in
        elif row.time_out:
            direction, time_val = 1, row.time_out
        else:
            continue

        emp_id = row.emp_code[1:].lstrip("0") or "0"
        emp = (
            ZucchettiEmployee.objects
            .filter(emp_code=emp_id.zfill(7))
            .first()
        )
        if not emp:
            continue

        card_code_13 = emp.card_code[2:]
        date_str = row.date.strftime("%d/%m/%Y")
        time_str = time_val.strftime("%H:%M")
        lines.append(f"{terminal}|{direction} 01 {card_code_13} {date_str} {time_str} 000000")

    if not lines:
        messages.error(request, "No lines generated.")
        return redirect("zucchetti_sync")

    content = "\r\n".join(lines) + "\r\n"
    filename = f"{rows.first().date.strftime('%Y.%m.%d')}.dat"
    response = HttpResponse(content, content_type="text/plain")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@zucchetti_required
def sync_view(request: HttpRequest) -> HttpResponse:
    active_tab = request.GET.get("tab", "import_zucchetti")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "import_zucchetti":
            xls_file = request.FILES.get("file")
            if xls_file:
                records = _parse_zucchetti_excel(xls_file)
                ZucchettiEmployee.objects.all().delete()
                ZucchettiEmployee.objects.bulk_create(records)
                messages.success(request, f"Imported {len(records)} Zucchetti employee records.")
            return redirect(f"{request.path}?tab=import_zucchetti")

        elif action == "clear_zucchetti":
            ZucchettiEmployee.objects.all().delete()
            messages.success(request, "Zucchetti employee list cleared.")
            return redirect(f"{request.path}?tab=import_zucchetti")

        elif action == "clear_request":
            NewRequest.objects.all().delete()
            messages.success(request, "New request list cleared.")
            return redirect(f"{request.path}?tab=new_request")

        elif action == "import_request":
            xls_file = request.FILES.get("file")
            if xls_file:
                records = _parse_request_excel(xls_file)
                NewRequest.objects.all().delete()
                NewRequest.objects.bulk_create(records)
                messages.success(request, f"Imported {len(records)} request rows.")
            return redirect(f"{request.path}?tab=new_request")

        elif action == "save_plant_settings":
            plants = request.POST.getlist("plant")
            terminals = request.POST.getlist("terminal")
            PlantSetting.objects.all().delete()
            PlantSetting.objects.bulk_create([
                PlantSetting(plant=p.strip(), terminal=t.strip())
                for p, t in zip(plants, terminals)
                if p.strip()
            ])
            messages.success(request, "Plant settings saved.")
            return redirect(f"{request.path}?tab=plant_settings")

        elif action == "check_dat":
            active_tab = "generate_csv"

    zucchetti_employees = ZucchettiEmployee.objects.all()
    request_rows = NewRequest.objects.all()
    plant_settings = PlantSetting.objects.all()

    dat_preview = None
    if active_tab == "generate_csv" and request.method == "POST" and request.POST.get("action") == "check_dat":
        dat_preview = _build_dat_preview()

    return render(request, "zucchetti/sync.html", {
        "zucchetti_employees": zucchetti_employees,
        "total_zucchetti": zucchetti_employees.count(),
        "request_rows": request_rows,
        "plant_settings": plant_settings,
        "active_tab": active_tab,
        "dat_preview": dat_preview,
    })
