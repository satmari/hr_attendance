from __future__ import annotations

import calendar
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    ExcelImportLog,
    ImportedData,
    PersonnelPresence,
    PersonnelPresenceDepartment,
    PersonnelMasterRecord,
    PersonnelActualData,
    PersonnelPresenceInteos,
    PersonnelPresenceLocation,
    PersonnelPresenceShift,
    PersonnelPresenceZucchetti,
)

MONTH_NAME_MAP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "januar": 1,
    "februar": 2,
    "mart": 3,
    "april": 4,
    "maj": 5,
    "jun": 6,
    "jul": 7,
    "avgust": 8,
    "septembar": 9,
    "oktobar": 10,
    "novembar": 11,
    "decembar": 12,
}

MAIN_FIELD_MAP = {
    "R": "employee_code",
    "NAME": "employee_name",
    "DEP sada": "current_department",
    "TYPE": "employee_type",
    "SUBSID sada": "current_location",
    "datum promene": "change_date",
    "DEP pre": "previous_department",
    "SUBSID pre": "previous_location",
    "START": "start_day",
    "END": "end_day",
    "KI TO SE": "ki_to_se",
    "ACTUAL": "actual",
}

MAT_FIELD_MAP = {
    "R": "employee_code",
    "START": "hire_date",
    "END": "termination_date",
    "RM": "job_title",
    "POGON": "department",
    "STATUS": "employment_status",
    "SENIORITY (m)": "seniority_months",
    "CATEGORY": "category",
}



@dataclass
class ParsedWorkbook:
    month: int
    year: int
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _detect_period(cells: list[Any]) -> tuple[int, int]:
    # First, try to find a direct datetime/date object
    for value in cells:
        if isinstance(value, (datetime, date)):
            return value.month, value.year

    # Then, try to parse string representations
    for value in cells:
        if value is None:
            continue

        text = str(value).strip()
        lower_text = text.lower()

        # Try "MM/YYYY", "MM-YYYY", "MM.YYYY"
        slash_match = re.search(r"(?P<month>\d{1,2})\s*[/.-]\s*(?P<year>\d{4})", text)
        if slash_match:
            return int(slash_match.group("month")), int(slash_match.group("year"))

        # Try "MonthName YYYY" or "YYYY MonthName"
        month_names_pattern = "|".join(re.escape(name) for name in MONTH_NAME_MAP.keys())
        
        # Pattern for "MonthName YYYY" or "YYYY MonthName"
        combined_month_year_pattern = rf"(?:(?P<month_name1>{month_names_pattern})\s*(?P<year1>\d{{4}}))|(?P<year2>\d{{4}})\s*(?P<month_name2>{month_names_pattern})"

        month_year_match = re.search(combined_month_year_pattern, lower_text)
        if month_year_match:
            year = int(month_year_match.group("year1") or month_year_match.group("year2"))
            matched_month_name = month_year_match.group("month_name1") or month_year_match.group("month_name2")
            for month_name, month_number in MONTH_NAME_MAP.items():
                if month_name == matched_month_name:
                    return month_number, year

    # Ako ne detektuje, vraća trenutni mesec i godinu kao sugestiju
    return datetime.now().month, datetime.now().year


def _load_workbook_from_bytes(file_bytes: bytes) -> Any:
    """Učitava Excel fajl iz bajtova u memoriji."""
    stream = BytesIO(file_bytes)
    return load_workbook(stream, data_only=True)

def parse_workbook(
    file_bytes: bytes,
    sheet_name: str,
    force_month: int | None = None,
    force_year: int | None = None,
    header_row: int = 6,
    data_start_row: int = 7,
) -> ParsedWorkbook:
    workbook = _load_workbook_from_bytes(file_bytes)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' was not found.")

    sheet = workbook[sheet_name]

    if force_month and force_year:
        month, year = force_month, force_year
    else:
        period_row = [cell.value for cell in sheet[5]]
        try:
            month, year = _detect_period(period_row)
        except ValueError:
            month, year = datetime.now().month, datetime.now().year

    headers = [_normalize_header(cell.value) for cell in sheet[header_row]]

    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        row_payload: dict[str, Any] = {}
        for column_index in range(len(headers)):
            h = headers[column_index]
            if h and h not in row_payload:  # first occurrence wins for duplicate headers
                row_payload[h] = row[column_index] if column_index < len(row) else None

        if not any(value not in ("", None) for value in row_payload.values()):
            continue

        row_payload["_row_number"] = row_index
        rows.append(row_payload)

    return ParsedWorkbook(month=month, year=year, sheet_name=sheet_name, headers=headers, rows=rows)


def _validate_sheets(workbook: Any) -> list[dict]:
    """
    Validates sheets in the workbook and returns a list of validation dicts.
    """
    sheet_specs = [
        {
            "name": "main",
            "required": True,
            "header_row": 6,
            "data_start_row": 7,
            "key_columns": ["R", "NAME", "DEP sada", "SUBSID sada", "KI TO SE", "START", "END"],
            "check_daily": True,
        },
        {
            "name": "mat",
            "required": False,
            "header_row": 2,
            "data_start_row": 3,
            "key_columns": ["R", "START", "END", "RM", "POGON", "STATUS"],
            "check_daily": False,
        },
        {
            "name": "ACTUAL data",
            "required": False,
            "header_row": 1,
            "data_start_row": 2,
            "key_columns": ["R", "IME"],
            "check_daily": False,
        },
    ]

    results = []
    for spec in sheet_specs:
        sname = spec["name"]
        found = sname in workbook.sheetnames

        if not found:
            results.append({
                "name": sname,
                "found": False,
                "required": spec["required"],
                "row_count": 0,
                "columns_ok": [],
                "columns_missing": spec["key_columns"][:],
                "has_daily_data": False,
            })
            continue

        ws = workbook[sname]
        header_row_idx = spec["header_row"]
        data_start_row = spec["data_start_row"]

        # Read headers from the specified header row
        headers_raw = [_normalize_header(cell.value) for cell in ws[header_row_idx]]
        headers_upper = [h.upper() for h in headers_raw]

        # Check key columns (case-insensitive exact match)
        columns_ok = []
        columns_missing = []
        for col in spec["key_columns"]:
            col_upper = col.upper()
            if col_upper in headers_upper:
                columns_ok.append(col)
            else:
                columns_missing.append(col)

        # Check for daily pattern columns (main sheet only)
        has_daily_data = False
        if spec["check_daily"]:
            has_dep_sub_sh = any(
                h.startswith("DEP") or h.startswith("SUB") or h.startswith("SH")
                for h in headers_upper
            )
            has_numeric = any(
                h.isdigit() and 1 <= int(h) <= 31
                for h in headers_upper
                if h.isdigit()
            )
            has_daily_data = has_dep_sub_sh and has_numeric

        # Count non-empty data rows
        row_count = 0
        for r in ws.iter_rows(min_row=data_start_row, values_only=True):
            if any(c is not None for c in r):
                row_count += 1

        results.append({
            "name": sname,
            "found": True,
            "required": spec["required"],
            "row_count": row_count,
            "columns_ok": columns_ok,
            "columns_missing": columns_missing,
            "has_daily_data": has_daily_data,
        })

    return results


def preview_workbook(file_bytes: bytes, limit: int = 10) -> dict[str, Any]:
    """
    Prikazuje preview isključivo za 'main' sheet.
    """
    workbook = _load_workbook_from_bytes(file_bytes)
    parsed = parse_workbook(file_bytes, sheet_name="main")

    extra_info = {}
    for sn in ["mat", "ACTUAL data"]:
        if sn in workbook.sheetnames:
            ws = workbook[sn]
            # Brzi brojač redova od reda 7 pa na dalje koji nisu skroz prazni
            count = 0
            for r in ws.iter_rows(min_row=7, values_only=True):
                if any(c is not None for c in r):
                    count += 1
            extra_info[sn] = count

    sheet_validation = _validate_sheets(workbook)

    return {
        "month": parsed.month,
        "year": parsed.year,
        "sheet_name": parsed.sheet_name,
        "headers": parsed.headers,
        "preview_rows": [
            {
                key: _cell_to_string(value)
                for key, value in row.items()
                if key != "_row_number"
            }
            for row in parsed.rows[:limit]
        ],
        "total_rows": len(parsed.rows),
        "extra_sheets_found": extra_info,
        "sheet_validation": sheet_validation,
    }


def _build_daily_payload(row: dict[str, Any], prefix: str) -> dict[str, str]:
    payload: dict[str, str] = {}
    for day in range(1, 32):
        payload[f"day{day}"] = _cell_to_string(row.get(f"{prefix}{day}", ""))
    return payload


def _build_inteos_payload(row: dict[str, Any]) -> dict[str, str]:
    payload: dict[str, str] = {}
    for day in range(1, 32):
        value = (
            row.get(f"{day}.")
            or row.get(f"{day} .")
            or row.get(f"{day}.-")
            or row.get(f"{day}.-{day}.")
            or row.get(f"{day}.-31.")
            or row.get(f"{day}.-31")
            or row.get(str(day))
            or ""
        )
        payload[f"day{day}"] = _cell_to_string(value)
    return payload


def _build_zucchetti_payload(row: dict[str, Any]) -> dict[str, str]:
    payload: dict[str, str] = {}
    for day in range(1, 32):
        value = row.get(str(day)) or row.get(f"{day} ") or row.get(f"{day}-31") or ""
        payload[f"day{day}"] = _cell_to_string(value)
    return payload


def _extract_main_payload(row: dict[str, Any], month: int, year: int) -> dict[str, str | int]:
    payload: dict[str, str | int] = {"month": month, "year": year}
    for header, field_name in MAIN_FIELD_MAP.items():
        payload[field_name] = _cell_to_string(row.get(header, ""))
    return payload


def _extract_mapped_payload(row: dict[str, Any], field_map: dict[str, str], month: int, year: int) -> dict[str, Any]:
    payload: dict[str, Any] = {"month": month, "year": year}
    for excel_header, model_field in field_map.items():
        found_key = next((k for k in row.keys() if excel_header.lower() in k.lower()), None)
        if found_key:
            payload[model_field] = _cell_to_string(row.get(found_key, ""))
    return payload


def _extract_exact_payload(row: dict[str, Any], field_map: dict[str, str], month: int, year: int) -> dict[str, Any]:
    """Exact (case-insensitive) header matching — avoids substring false matches like 'R' → 'NR'."""
    payload: dict[str, Any] = {"month": month, "year": year}
    row_upper = {k.strip().upper(): v for k, v in row.items()}
    for excel_header, model_field in field_map.items():
        val = row_upper.get(excel_header.strip().upper())
        if val is not None:
            payload[model_field] = _cell_to_string(val)
    return payload


def _upsert_daily_model(model_class, presence: PersonnelPresence, values: dict[str, str]) -> None:
    defaults = {"presence": presence, **values}
    model_class.objects.update_or_create(presence=presence, defaults=defaults)


@transaction.atomic
def delete_period_data(month: int, year: int):
    """Briše apsolutno sve podatke vezane za određeni mesec i godinu."""
    PersonnelPresence.objects.filter(month=month, year=year).delete()
    PersonnelMasterRecord.objects.filter(month=month, year=year).delete()
    PersonnelActualData.objects.filter(month=month, year=year).delete()
    ExcelImportLog.objects.filter(month=month, year=year).delete()


@transaction.atomic
def import_workbook(
    file_bytes: bytes,
    file_name: str,
    *,
    selected_month: int | None = None,
    selected_year: int | None = None,
    replace_existing: bool = False,
) -> ExcelImportLog:
    main_parsed = parse_workbook(file_bytes, sheet_name="main", force_month=selected_month, force_year=selected_year)
    month = selected_month or main_parsed.month
    year = selected_year or main_parsed.year

    import_log = ExcelImportLog.objects.create(
        original_file_name=file_name,
        month=month,
        year=year,
        status=ExcelImportLog.STATUS_PENDING,
        replace_existing=replace_existing,
        detected_sheet="multiple",
        total_rows=len(main_parsed.rows),
        message="Import started (Main, Mat, Actual).",
    )

    if replace_existing:
        delete_period_data(month, year)

    # 1. Import MAIN sheeta
    for row in main_parsed.rows:
        ImportedData.objects.create(
            import_log=import_log,
            row_number=row["_row_number"],
            raw_payload={key: _cell_to_string(value) for key, value in row.items() if key != "_row_number"},
        )

        main_payload = _extract_main_payload(row, month, year)
        employee_code = str(main_payload.get("employee_code", "")).strip()
        if not employee_code:
            continue

        presence, _created = PersonnelPresence.objects.update_or_create(
            employee_code=employee_code,
            month=month,
            year=year,
            defaults=main_payload,
        )

        _upsert_daily_model(PersonnelPresenceDepartment, presence, _build_daily_payload(row, "DEP"))
        _upsert_daily_model(PersonnelPresenceLocation, presence, _build_daily_payload(row, "SUB"))
        _upsert_daily_model(PersonnelPresenceShift, presence, _build_daily_payload(row, "SH"))
        _upsert_daily_model(PersonnelPresenceInteos, presence, _build_inteos_payload(row))
        _upsert_daily_model(PersonnelPresenceZucchetti, presence, _build_zucchetti_payload(row))

    # 2. Import MAT sheeta (ako postoji) — zaglavlja su na redu 2, podaci od reda 3
    workbook = _load_workbook_from_bytes(file_bytes)
    if "mat" in workbook.sheetnames:
        mat_parsed = parse_workbook(
            file_bytes, sheet_name="mat",
            force_month=month, force_year=year,
            header_row=2, data_start_row=3,
        )
        for row in mat_parsed.rows:
            payload = _extract_exact_payload(row, MAT_FIELD_MAP, month, year)
            if payload.get("employee_code"):
                PersonnelMasterRecord.objects.update_or_create(
                    employee_code=payload["employee_code"],
                    month=month,
                    year=year,
                    defaults=payload,
                )

    # 3. Import ACTUAL data sheeta (ako postoji) — kolone A:D (R, IME, actual_pos_after, actual_pos_before)
    if "ACTUAL data" in workbook.sheetnames:
        act_parsed = parse_workbook(file_bytes, sheet_name="ACTUAL data", force_month=month, force_year=year,
                                    header_row=1, data_start_row=2)
        # Detektuj dve "ACTUAL pos." kolone po redosledu pojavljivanja u headerima
        actual_pos_cols = [h for h in act_parsed.headers if h.lower().startswith("actual pos")]
        col_after = actual_pos_cols[0] if len(actual_pos_cols) >= 1 else None
        col_before = actual_pos_cols[1] if len(actual_pos_cols) >= 2 else None

        for row in act_parsed.rows:
            employee_code = _cell_to_string(row.get("R", "")).strip()
            if not employee_code:
                continue
            PersonnelActualData.objects.update_or_create(
                employee_code=employee_code,
                month=month,
                year=year,
                defaults={
                    "employee_name": _cell_to_string(row.get("IME", "")),
                    "actual_pos_after": _cell_to_string(row.get(col_after, "")) if col_after else "",
                    "actual_pos_before": _cell_to_string(row.get(col_before, "")) if col_before else "",
                },
            )

    import_log.imported_rows = PersonnelPresence.objects.filter(month=month, year=year).count()
    import_log.status = ExcelImportLog.STATUS_COMPLETED
    import_log.message = f"Import completed for {calendar.month_name[month]} {year}."
    import_log.save(update_fields=["imported_rows", "status", "message", "updated_at"])
    return import_log
